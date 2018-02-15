# -*- coding: utf-8 -*-
#------------------------------------------------------#
# ExcelDiff                                            #
# Comparison of the differences between two excel      #
# and light the row, the column and the cell in detail #
# by YuLong Tu 2018/2/7                               #
#------------------------------------------------------#

from PyQt4 import QtCore, QtGui
import icon,logo# 装载qrc
import xlrd
import openpyxl
import sys

# 全局变量声明
workBook1=[]# 旧excel
workBook2=[]# 新excel
dicRowAdd={}# 行增
dicColAdd={}# 列增
dicRowDel={}# 行删
dicColDel={}# 列删
dicCellAdd={}# 单元格改变
dicCellDel={}# 单元格改变
# 记录数据在对比表格中的实际位置,即原本位置+偏移
dicRowAdd1={}
dicColAdd1={}
dicRowDel1={}
dicColDel1={}
showTool=0 # 工具栏状态
flagOld=0# 文件格式为xls
flagNew=0#
global sheet1,sheet2
global nrows1,ncols1,nrows2,ncols2
global excelWindow1,excelWindow2,excelWindow,excelMain

try:
    _fromUtf8 = QtCore.QString.fromUtf8
except AttributeError:
    def _fromUtf8(s):
        return s

try:
    _encoding = QtGui.QApplication.UnicodeUTF8
    def _translate(context, text, disambig):
        return QtGui.QApplication.translate(context, text, disambig, _encoding)
except AttributeError:
    def _translate(context, text, disambig):
        return QtGui.QApplication.translate(context, text, disambig)

# 登录验证框
class Ui_Log(object):
    def setupUi(self, Log):
        Log.setObjectName(_fromUtf8("Log"))
        Log.setEnabled(True)
        Log.resize(506, 201)
        Log.setWindowIcon(QtGui.QIcon(':/ico/Cut.ico'))
        Log.setStyleSheet(_fromUtf8("background-color: rgb(255, 255, 255);"))
        self.centralwidget = QtGui.QWidget(Log)
        self.centralwidget.setObjectName(_fromUtf8("centralwidget"))
        self.layoutWidget = QtGui.QWidget(self.centralwidget)
        self.layoutWidget.setGeometry(QtCore.QRect(60, 60, 391, 81))
        self.layoutWidget.setObjectName(_fromUtf8("layoutWidget"))
        self.gridLayout = QtGui.QGridLayout(self.layoutWidget)
        self.gridLayout.setObjectName(_fromUtf8("gridLayout"))
        self.pushButton_Login = QtGui.QPushButton(self.layoutWidget)
        self.pushButton_Login.setEnabled(False)
        self.pushButton_Login.setObjectName(_fromUtf8("pushButton_Login"))
        self.gridLayout.addWidget(self.pushButton_Login, 0, 2, 1, 1)
        self.pushButton_Clear = QtGui.QPushButton(self.layoutWidget)
        self.pushButton_Clear.setEnabled(True)
        self.pushButton_Clear.setObjectName(_fromUtf8("pushButton_Clear"))
        self.gridLayout.addWidget(self.pushButton_Clear, 1, 2, 1, 1)
        self.label_UserName = QtGui.QLabel(self.layoutWidget)
        self.label_UserName.setObjectName(_fromUtf8("label_UserName"))
        self.gridLayout.addWidget(self.label_UserName, 0, 0, 1, 1)
        self.label_Password = QtGui.QLabel(self.layoutWidget)
        self.label_Password.setObjectName(_fromUtf8("label_Password"))
        self.gridLayout.addWidget(self.label_Password, 1, 0, 1, 1)
        self.lineEdit = QtGui.QLineEdit(self.layoutWidget)
        self.lineEdit.setStyleSheet(_fromUtf8(""))
        self.lineEdit.setObjectName(_fromUtf8("lineEdit"))
        self.lineEdit.setContextMenuPolicy(QtCore.Qt.NoContextMenu)  # 关闭右键菜单
        self.gridLayout.addWidget(self.lineEdit, 0, 1, 1, 1)
        self.lineEdit_2 = QtGui.QLineEdit(self.layoutWidget)
        self.lineEdit_2.setText(_fromUtf8(""))
        self.lineEdit_2.setObjectName(_fromUtf8("lineEdit_2"))
        self.lineEdit_2.setEchoMode((QtGui.QLineEdit.Password))# 密码显示
        self.lineEdit_2.setContextMenuPolicy(QtCore.Qt.NoContextMenu)# 关闭右键菜单
        self.gridLayout.addWidget(self.lineEdit_2, 1, 1, 1, 1)
        Log.setCentralWidget(self.centralwidget)
        self.statusbar = QtGui.QStatusBar(Log)
        self.statusbar.setObjectName(_fromUtf8("statusbar"))
        Log.setStatusBar(self.statusbar)
        self.label_UserName.setBuddy(self.lineEdit)
        self.label_Password.setBuddy(self.lineEdit_2)

        self.retranslateUi(Log)

        QtCore.QObject.connect(self.lineEdit, QtCore.SIGNAL(_fromUtf8("textChanged(QString)")),self.enableLoginButton)
        QtCore.QObject.connect(self.lineEdit_2, QtCore.SIGNAL(_fromUtf8("textChanged(QString)")), self.enableLoginButton)
        QtCore.QObject.connect(self.pushButton_Login, QtCore.SIGNAL(_fromUtf8("clicked()")), self.checkOut)
        QtCore.QObject.connect(self.pushButton_Clear, QtCore.SIGNAL(_fromUtf8("clicked()")), self.lineEdit.clear)
        QtCore.QObject.connect(self.pushButton_Clear, QtCore.SIGNAL(_fromUtf8("clicked()")), self.lineEdit_2.clear)
        QtCore.QMetaObject.connectSlotsByName(Log)

        Log.setTabOrder(self.lineEdit, self.lineEdit_2)
        Log.setTabOrder(self.lineEdit_2, self.pushButton_Login)
        Log.setTabOrder(self.pushButton_Login, self.pushButton_Clear)
        # 延迟启动
        QtCore.QThread.sleep(2)

    def retranslateUi(self, Log):
        Log.setWindowTitle(_translate("Log", "Log", None))
        self.pushButton_Login.setText(_translate("Log", "&Login", None))
        self.pushButton_Clear.setText(_translate("Log", "&Clear", None))
        self.label_UserName.setText(_translate("Log", "&UserName:", None))
        self.label_Password.setText(_translate("Log", "&Password:", None))

        file = QtCore.QFile('csstest.qss')
        file.open(QtCore.QFile.ReadOnly)
        styleSheet = file.readAll()
        styleSheet = unicode(styleSheet, encoding='utf8')
        QtGui.qApp.setStyleSheet(styleSheet)
    # 只有账号密码都填写了Login按钮才会点亮
    def enableLoginButton(self):
        if self.lineEdit.text()!='' and self.lineEdit_2.text()!='':
            self.pushButton_Login.setEnabled(True)
        else:
            # 当按钮为亮时，变为灰色
            if self.pushButton_Login.autoDefault()!=True:
                self.pushButton_Login.setEnabled(False)

    def checkOut(self):
        user_name = self.lineEdit.text()
        pass_word = self.lineEdit_2.text()
        # 仅此一个账号密码，本来想加入数据库内容，但是查到的资料都说最后无法将数据库内容一起打包为exe，遂放弃
        if user_name == "tyl" and pass_word == "tyl":
            logWindow.close()
            mainWindow.show()
        else:
            reply = QtGui.QMessageBox.warning(self.centralwidget,'Message',"Sorry, the account with this keycode was not found.")
            self.lineEdit.clear()
            self.lineEdit_2.clear()
    # python有时候会崩溃，还不清楚什么原因

# 主窗口
class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName(_fromUtf8("MainWindow"))
        MainWindow.resize(619, 557)
        MainWindow.setFixedSize(MainWindow.width(), MainWindow.height());
        MainWindow.setStyleSheet(_fromUtf8("background-color: rgb(255, 255, 255);"))
        MainWindow.setWindowIcon(QtGui.QIcon(':/ico/Cut.ico'))
        self.centralwidget = QtGui.QWidget(MainWindow)
        self.centralwidget.setObjectName(_fromUtf8("centralwidget"))
        self.label_Logo = QtGui.QLabel(self.centralwidget)
        self.label_Logo.setGeometry(QtCore.QRect(410, 370, 161, 101))
        self.label_Logo.setStyleSheet(_fromUtf8("image: url(:/ico/timg.jpg);"))
        self.label_Logo.setText(_fromUtf8(""))
        self.label_Logo.setObjectName(_fromUtf8("label_Logo"))
        self.label_OldVersionFile = QtGui.QLabel(self.centralwidget)
        self.label_OldVersionFile.setGeometry(QtCore.QRect(30, 50, 121, 16))
        self.label_OldVersionFile.setObjectName(_fromUtf8("label_OldVersionFile"))
        self.label_NewVersionFile = QtGui.QLabel(self.centralwidget)
        self.label_NewVersionFile.setGeometry(QtCore.QRect(30, 90, 121, 16))
        self.label_NewVersionFile.setObjectName(_fromUtf8("label_NewVersionFile"))
        self.toolButton_Old = QtGui.QToolButton(self.centralwidget)
        self.toolButton_Old.setGeometry(QtCore.QRect(550, 50, 31, 21))
        self.toolButton_Old.setObjectName(_fromUtf8("toolButton_Old"))
        self.comboBox_Old = QtGui.QComboBox(self.centralwidget)
        self.comboBox_Old.setGeometry(QtCore.QRect(170, 50, 371, 22))
        self.comboBox_Old.setObjectName(_fromUtf8("comboBox_Old"))
        self.toolButton_New = QtGui.QToolButton(self.centralwidget)
        self.toolButton_New.setGeometry(QtCore.QRect(550, 90, 31, 21))
        self.toolButton_New.setObjectName(_fromUtf8("toolButton_New"))
        self.comboBox_New = QtGui.QComboBox(self.centralwidget)
        self.comboBox_New.setGeometry(QtCore.QRect(170, 90, 371, 22))
        self.comboBox_New.setObjectName(_fromUtf8("comboBox_New"))
        self.tabWidget_All = QtGui.QTabWidget(self.centralwidget)
        self.tabWidget_All.setGeometry(QtCore.QRect(30, 170, 551, 211))
        self.tabWidget_All.setTabPosition(QtGui.QTabWidget.West)
        self.tabWidget_All.setObjectName(_fromUtf8("tabWidget_All"))
        self.tab_Add = QtGui.QWidget()
        self.tab_Add.setObjectName(_fromUtf8("tab_Add"))
        self.tabWidget_Add = QtGui.QTabWidget(self.tab_Add)
        self.tabWidget_Add.setGeometry(QtCore.QRect(0, 0, 531, 211))
        self.tabWidget_Add.setTabPosition(QtGui.QTabWidget.North)
        self.tabWidget_Add.setObjectName(_fromUtf8("tabWidget_Add"))
        self.tab_ColumnChange_2 = QtGui.QWidget()
        self.tab_ColumnChange_2.setObjectName(_fromUtf8("tab_ColumnChange_2"))

        self.listWidget_Add_Col = QtGui.QListWidget(self.tab_ColumnChange_2)
        self.listWidget_Add_Col.setGeometry(QtCore.QRect(0, 0, 531, 191))
        self.listWidget_Add_Col.setObjectName(_fromUtf8("listWidget_Add_Col"))

        self.tabWidget_Add.addTab(self.tab_ColumnChange_2, _fromUtf8(""))
        self.tab_RowChange_2 = QtGui.QWidget()
        self.tab_RowChange_2.setObjectName(_fromUtf8("tab_RowChange_2"))

        self.listWidget__Add_Row = QtGui.QListWidget(self.tab_RowChange_2)
        self.listWidget__Add_Row.setGeometry(QtCore.QRect(0, 0, 531, 191))
        self.listWidget__Add_Row.setObjectName(_fromUtf8("listWidget__Add_Row"))

        self.tabWidget_Add.addTab(self.tab_RowChange_2, _fromUtf8(""))
        self.tab_CellChange_2 = QtGui.QWidget()
        self.tab_CellChange_2.setObjectName(_fromUtf8("tab_CellChange_2"))

        self.listWidget__Add_Cell = QtGui.QListWidget(self.tab_CellChange_2)
        self.listWidget__Add_Cell.setGeometry(QtCore.QRect(0, 0, 531, 191))
        self.listWidget__Add_Cell.setObjectName(_fromUtf8("listWidget__Add_Cell"))

        self.tabWidget_Add.addTab(self.tab_CellChange_2, _fromUtf8(""))
        self.tabWidget_All.addTab(self.tab_Add, _fromUtf8(""))
        self.tab_Delete = QtGui.QWidget()
        self.tab_Delete.setObjectName(_fromUtf8("tab_Delete"))
        self.tabWidget_Delete = QtGui.QTabWidget(self.tab_Delete)
        self.tabWidget_Delete.setGeometry(QtCore.QRect(0, 0, 541, 211))
        self.tabWidget_Delete.setTabPosition(QtGui.QTabWidget.North)
        self.tabWidget_Delete.setObjectName(_fromUtf8("tabWidget_Delete"))
        self.tab_ColumnChange = QtGui.QWidget()
        self.tab_ColumnChange.setObjectName(_fromUtf8("tab_ColumnChange"))

        self.listWidget_Del_Col_2 = QtGui.QListWidget(self.tab_ColumnChange)
        self.listWidget_Del_Col_2.setGeometry(QtCore.QRect(0, 0, 531, 191))
        self.listWidget_Del_Col_2.setObjectName(_fromUtf8("listWidget_Del_Col_2"))

        self.tabWidget_Delete.addTab(self.tab_ColumnChange, _fromUtf8(""))
        self.tab_RowChange = QtGui.QWidget()
        self.tab_RowChange.setObjectName(_fromUtf8("tab_RowChange"))

        self.listWidget_Del_Row = QtGui.QListWidget(self.tab_RowChange)
        self.listWidget_Del_Row.setGeometry(QtCore.QRect(0, 0, 531, 191))
        self.listWidget_Del_Row.setObjectName(_fromUtf8("listWidget_Del_Row"))

        self.tabWidget_Delete.addTab(self.tab_RowChange, _fromUtf8(""))
        self.tab_CellChange = QtGui.QWidget()
        self.tab_CellChange.setObjectName(_fromUtf8("tab_CellChange"))

        self.listWidget_Del_Cell = QtGui.QListWidget(self.tab_CellChange)
        self.listWidget_Del_Cell.setGeometry(QtCore.QRect(0, 0, 531, 191))
        self.listWidget_Del_Cell.setObjectName(_fromUtf8("listWidget_Del_Cell"))

        self.tabWidget_Delete.addTab(self.tab_CellChange, _fromUtf8(""))
        self.tabWidget_All.addTab(self.tab_Delete, _fromUtf8(""))
        self.pushButton_Start = QtGui.QPushButton(self.centralwidget)
        self.pushButton_Start.setGeometry(QtCore.QRect(170, 420, 93, 28))
        self.pushButton_Start.setObjectName(_fromUtf8("pushButton_Start"))
        self.label = QtGui.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(30, 130, 241, 16))
        self.label.setObjectName(_fromUtf8("label"))
        self.comboBox = QtGui.QComboBox(self.centralwidget)
        self.comboBox.setGeometry(QtCore.QRect(130, 130, 87, 22))
        self.comboBox.setEditable(False)
        self.comboBox.setObjectName(_fromUtf8("comboBox"))
        self.pushButton_Load = QtGui.QPushButton(self.centralwidget)
        self.pushButton_Load.setGeometry(QtCore.QRect(60, 420, 93, 28))
        self.pushButton_Load.setObjectName(_fromUtf8("pushButton_Load"))

        self.label_2 = QtGui.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(230, 130, 71, 20))
        self.label_2.setObjectName(_fromUtf8("label_2"))
        self.comboBox_2 = QtGui.QComboBox(self.centralwidget)
        self.comboBox_2.setGeometry(QtCore.QRect(300, 130, 87, 22))
        self.comboBox_2.setEditable(False)
        self.comboBox_2.setObjectName(_fromUtf8("comboBox_2"))
        self.label_3 = QtGui.QLabel(self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(400, 130, 71, 21))
        self.label_3.setObjectName(_fromUtf8("label_3"))
        self.comboBox_3 = QtGui.QComboBox(self.centralwidget)
        self.comboBox_3.setGeometry(QtCore.QRect(460, 130, 87, 22))
        self.comboBox_3.setEditable(False)
        self.comboBox_3.setObjectName(_fromUtf8("comboBox_3"))

        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtGui.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 619, 26))
        self.menubar.setObjectName(_fromUtf8("menubar"))
        self.menu = QtGui.QMenu(self.menubar)
        self.menu.setObjectName(_fromUtf8("menu"))
        self.menu_2 = QtGui.QMenu(self.menubar)
        self.menu_2.setObjectName(_fromUtf8("menu_2"))
        self.menu_3 = QtGui.QMenu(self.menubar)
        self.menu_3.setObjectName(_fromUtf8("menu_3"))
        MainWindow.setMenuBar(self.menubar)
        self.toolBar = QtGui.QToolBar(MainWindow)
        self.toolBar.setObjectName(_fromUtf8("toolBar"))
        MainWindow.addToolBar(QtCore.Qt.TopToolBarArea, self.toolBar)
        self.toolBar.setVisible(False)
        self.actionTool_Bar = QtGui.QAction(MainWindow)
        self.actionTool_Bar.setCheckable(True)
        self.actionTool_Bar.setObjectName(_fromUtf8("actionTool_Bar"))
        self.action_About = QtGui.QAction(MainWindow)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap(_fromUtf8(":/ico/About.ico")), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.action_About.setIcon(icon)
        self.action_About.setObjectName(_fromUtf8("action_About"))
        self.action_Save = QtGui.QAction(MainWindow)
        icon1 = QtGui.QIcon()
        icon1.addPixmap(QtGui.QPixmap(_fromUtf8(":/ico/Save.ico")), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.action_Save.setIcon(icon1)
        self.action_Save.setObjectName(_fromUtf8("action_Save"))
        #self.action_Print = QtGui.QAction(MainWindow)
        icon2 = QtGui.QIcon()
        icon2.addPixmap(QtGui.QPixmap(_fromUtf8(":/ico/Copy.ico")), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        #self.action_Print.setIcon(icon2)
        #self.action_Print.setObjectName(_fromUtf8("action_Print"))
        self.action_Quit = QtGui.QAction(MainWindow)
        self.action_Quit.setObjectName(_fromUtf8("action_Quit"))
        self.menu.addAction(self.action_Save)
        #self.menu.addAction(self.action_Print)
        self.menu.addAction(self.action_Quit)
        self.menu_2.addAction(self.actionTool_Bar)
        self.menu_3.addAction(self.action_About)
        self.menubar.addAction(self.menu.menuAction())
        self.menubar.addAction(self.menu_2.menuAction())
        self.menubar.addAction(self.menu_3.menuAction())
        self.toolBar.addAction(self.action_Save)
        #self.toolBar.addAction(self.action_Print)
        self.toolBar.addAction(self.action_About)

        self.retranslateUi(MainWindow)
        # 触发事件
        QtCore.QObject.connect(self.toolButton_Old, QtCore.SIGNAL(_fromUtf8("clicked()")), self.openFileOld)
        QtCore.QObject.connect(self.toolButton_New, QtCore.SIGNAL(_fromUtf8("clicked()")), self.openFileNew)
        QtCore.QObject.connect(self.pushButton_Load, QtCore.SIGNAL(_fromUtf8("clicked()")), self.startLoadExcel)
        QtCore.QObject.connect(self.pushButton_Start, QtCore.SIGNAL(_fromUtf8("clicked()")), self.startCompareExcel)
        QtCore.QObject.connect(self.pushButton_Start, QtCore.SIGNAL(_fromUtf8("clicked()")),self.printExcel)
        QtCore.QObject.connect(self.listWidget__Add_Cell,QtCore.SIGNAL(_fromUtf8("itemClicked(QListWidgetItem *)")),
                               self.showExcel)
        QtCore.QObject.connect(self.listWidget_Del_Cell, QtCore.SIGNAL(_fromUtf8("itemClicked(QListWidgetItem *)")),
                               self.showExcel)
        QtCore.QObject.connect(self.listWidget_Del_Col_2, QtCore.SIGNAL(_fromUtf8("itemClicked(QListWidgetItem *)")),
                               self.showExcel)
        QtCore.QObject.connect(self.listWidget__Add_Row, QtCore.SIGNAL(_fromUtf8("itemClicked(QListWidgetItem *)")),
                               self.showExcel)
        QtCore.QObject.connect(self.listWidget_Add_Col, QtCore.SIGNAL(_fromUtf8("itemClicked(QListWidgetItem *)")),
                               self.showExcel)
        QtCore.QObject.connect(self.listWidget_Del_Row, QtCore.SIGNAL(_fromUtf8("itemClicked(QListWidgetItem *)")),
                               self.showExcel)
        # 触发动作
        self.action_Quit.triggered.connect(QtGui.qApp.quit)
        self.action_About.triggered.connect(self.aboutLittleDiff)
        self.actionTool_Bar.triggered.connect(self.showToolBar)
        #self.action_Print.triggered.connect(self.printLog)
        self.action_Save.triggered.connect(self.saveExcel)

        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        self.tabWidget_All.setCurrentIndex(1)
        self.tabWidget_Add.setCurrentIndex(2)
        self.tabWidget_Delete.setCurrentIndex(2)

    def retranslateUi(self, MainWindow):
        MainWindow.setWindowTitle(_translate("MainWindow", "LittleDiff", None))
        self.label_OldVersionFile.setText(_translate("MainWindow", "OldVersionFile:", None))
        self.label_NewVersionFile.setText(_translate("MainWindow", "NewVersionFile:", None))
        self.toolButton_Old.setText(_translate("MainWindow", "...", None))
        self.toolButton_New.setText(_translate("MainWindow", "...", None))
        self.tabWidget_Add.setTabText(self.tabWidget_Add.indexOf(self.tab_ColumnChange_2), _translate("MainWindow", "ColumnChange", None))
        self.tabWidget_Add.setTabText(self.tabWidget_Add.indexOf(self.tab_RowChange_2), _translate("MainWindow", "RowChange", None))
        self.tabWidget_Add.setTabText(self.tabWidget_Add.indexOf(self.tab_CellChange_2), _translate("MainWindow", "CellChange", None))
        self.tabWidget_All.setTabText(self.tabWidget_All.indexOf(self.tab_Add), _translate("MainWindow", "Add", None))
        self.tabWidget_Delete.setTabText(self.tabWidget_Delete.indexOf(self.tab_ColumnChange), _translate("MainWindow", "ColumnChange", None))
        self.tabWidget_Delete.setTabText(self.tabWidget_Delete.indexOf(self.tab_RowChange), _translate("MainWindow", "RowChange", None))
        self.tabWidget_Delete.setTabText(self.tabWidget_Delete.indexOf(self.tab_CellChange), _translate("MainWindow", "CellChange", None))
        self.tabWidget_All.setTabText(self.tabWidget_All.indexOf(self.tab_Delete), _translate("MainWindow", "Delete", None))
        self.pushButton_Start.setText(_translate("MainWindow", "Start", None))
        self.label.setText(_translate("MainWindow", "Comparable：", None))
        self.pushButton_Load.setText(_translate("MainWindow", "Load", None))
        self.label_2.setText(_translate("MainWindow", "Deleted：", None))
        self.label_3.setText(_translate("MainWindow", "Added：", None))
        self.menu.setTitle(_translate("MainWindow", "File", None))
        self.menu_2.setTitle(_translate("MainWindow", "Tools", None))
        self.menu_3.setTitle(_translate("MainWindow", "Help", None))
        self.toolBar.setWindowTitle(_translate("MainWindow", "toolBar", None))
        self.actionTool_Bar.setText(_translate("MainWindow", "Tool Bar", None))
        self.actionTool_Bar.setToolTip(_translate("MainWindow", "Tool Bar", None))
        self.actionTool_Bar.setShortcut(_translate("MainWindow", "Ctrl+T", None))
        self.action_About.setText(_translate("MainWindow", "About LittleDiff", None))
        self.action_About.setToolTip(_translate("MainWindow", "About LittleDiff", None))
        self.action_About.setShortcut(_translate("MainWindow", "Ctrl+A", None))
        self.action_Save.setText(_translate("MainWindow", "Save", None))
        self.action_Save.setToolTip(_translate("MainWindow", "Save", None))
        self.action_Save.setShortcut(_translate("MainWindow", "Ctrl+S", None))
        #self.action_Print.setText(_translate("MainWindow", "Print", None))
        #self.action_Print.setShortcut(_translate("MainWindow", "Ctrl+P", None))
        self.action_Quit.setText(_translate("MainWindow", "Quit", None))
        self.action_Quit.setToolTip(_translate("MainWindow", "Quit", None))
        self.action_Quit.setShortcut(_translate("MainWindow", "Ctrl+Q", None))

    # 后来了解到可以通过sender()判断信号来源，openFileOld与openFileNew可以合并为一个函数
    def openFileOld(self):
        try:
            fileName = QtGui.QFileDialog.getOpenFileName(self.centralwidget, 'Open file', './',
                                                     'Excel Files (*.xls);;Excel Files (*.xlsx);;All Files (*)')
            if fileName != '':
                self.comboBox_Old.addItem(fileName)
            # 显示新加入的最后一项
            self.comboBox_Old.setCurrentIndex(self.comboBox_Old.count()-1)
        except IOError:
            return

    def openFileNew(self):
        try:
            fileName = QtGui.QFileDialog.getOpenFileName(self.centralwidget, 'Open file', './',
                                                     'Excel Files (*.xls);;Excel Files (*.xlsx);;All Files (*)')
            if fileName!='':
                self.comboBox_New.addItem(fileName)
            # 显示新加入的最后一项
            self.comboBox_New.setCurrentIndex(self.comboBox_New.count()-1)
        except IOError:
            return
    # 点击Load按钮触发
    def startLoadExcel(self):
        global workBook1, workBook2, flagNew,flagOld
        workBook1={}
        workBook2={}
        try:
            #xlrd.Book.encoding = "gbk"
            fileNameNew = unicode(self.comboBox_New.currentText())
            fileNameNewSplit = fileNameNew.split('.')[1]
            if fileNameNewSplit==u'xlsx':
                workBook2 = openpyxl.load_workbook(fileNameNew)
                names2 = workBook2.sheetnames
                flagNew = 1
            else:
                workBook2 =xlrd.open_workbook(fileNameNew)
                names2 = workBook2.sheet_names()
                flagNew=0

            fileNameOdd = unicode(self.comboBox_Old.currentText())
            fileNameOddSplit = fileNameOdd.split('.')[1]
            if fileNameOddSplit == u'xlsx':
                workBook1 = openpyxl.load_workbook(fileNameOdd)
                names1 = workBook1.sheetnames
                flagOld = 1
            else:
                workBook1 = xlrd.open_workbook(fileNameOdd)
                names1 = workBook1.sheet_names()
                flagOld=0

            intersection = []
            # 获取excel中同名sheet，但似乎求交集的方法会打乱列表名
            intersection = list(set(names1).intersection(set(names2)))
            #print intersection
            intersectiontep=[]

            for i,j in enumerate(intersection):
                content = []
                content02 = []
                if flagOld ==0:
                    sheet=workBook1.sheet_by_name(unicode(j))
                    nrows01 = sheet.nrows
                    ncols01 = sheet.ncols
                    #print nrows01, ncols01
                    for m in range(0,nrows01):
                        contentR=sheet.row_values(m)
                        for n in range(0,ncols01):
                            if unicode(contentR[n])!=u''and unicode(contentR[n])!=u' ':
                                content.append(unicode(contentR[n]))
                else:
                    sheet = workBook1[unicode(j)]
                    nrows01 = sheet.max_row
                    ncols01 = sheet.max_column

                    for m in range(0,nrows01):
                        for n in range(0,ncols01):
                            t=sheet.cell(row=m+1, column=n+1).value
                            if unicode(t)!=u'None'and unicode(t)!=u' ':
                                if (type(t)==type(1L) or type(t)==type(1)):
                                    content.append(unicode(float(sheet.cell(row=m + 1, column=n + 1).value)))
                                else:
                                    content.append(unicode((sheet.cell(row=m + 1, column=n + 1).value)))

                if flagNew ==0:
                    sheet02=workBook2.sheet_by_name(unicode(j))
                    nrows02 = sheet02.nrows
                    ncols02 = sheet02.ncols
                    for m in range(0,nrows02):
                        contentR=sheet02.row_values(m)
                        for n in range(0,ncols02):
                            if unicode(contentR[n]) != u'' and unicode(contentR[n]) != u' ':
                                content02.append(unicode(contentR[n]))
                else:
                    sheet02 = workBook2[unicode(j)]
                    nrows02 = sheet02.max_row
                    ncols02 = sheet02.max_column
                    #print nrows02,ncols02
                    for m in range(0, nrows02):
                        for n in range(0, ncols02):
                            t=sheet02.cell(row=m + 1, column=n + 1).value
                            #print type(t)
                            if unicode(t) != u'None'and unicode(t)!=u' ':
                                if type(t)==type(1L) or type(t)==type(1):
                                    content02.append(unicode(float(sheet02.cell(row=m + 1, column=n + 1).value)))
                                else:
                                    content02.append(unicode((sheet02.cell(row=m + 1, column=n + 1).value)))
                if content==content02:
                    intersectiontep.append(j)
            #print intersectiontep
            if intersectiontep==intersection and names1==names2:
                reply = QtGui.QMessageBox.warning(self.centralwidget, 'Message',
                                                  "Sorry, the content of excels is the same")
                self.comboBox.clear()
                self.comboBox_2.clear()
                self.comboBox_3.clear()
                return
            #print intersectiontep
            self.comboBox.clear()
            self.comboBox_2.clear()
            self.comboBox_3.clear()
            intersection1=list(set(intersection).difference(set(intersectiontep)))

            delet=list(set(names1).difference(set(intersection)))
            add=list(set(names2).difference(set(intersection)))

            #print names1
            #print names2
            #print intersection
            #print intersection1
            #print delet
            #print add
            if intersection1==[]:
                self.comboBox.addItem(u'无')
                #return
            else:
                for i in intersection1:
                    self.comboBox.addItem(i)
            for i in delet:
                self.comboBox_2.addItem(i)
            for i in add:
                self.comboBox_3.addItem(i)
        except (IOError,IndexError):
            return
    #
    def startCompareExcel(self):
        try:
            global workBook1, workBook2, sheet1, sheet2,flagNew,flagOld,nrows1,ncols1,nrows2,ncols2
            row1 = []
            column1 = []
            row2 = []
            column2 = []
            # 选中的同名表项
            nameTemp = unicode(self.comboBox.currentText())
            # 读取时需要判断是否为空sheet
            if flagOld==0:
                sheet1 = workBook1.sheet_by_name(nameTemp)
                try:
                    rowtemp1 = sheet1.col_values(0)
                    columntemp1 = sheet1.row_values(0)
                    for i in columntemp1:
                        column1.append(unicode(i))
                    for i in rowtemp1:
                        row1.append(unicode(i))
                except IndexError:
                    column1=[u'None']
                    row1=[u'None']



                #temp=SheetSpace(sheet1,flagOld)
                #if(temp[0]!=-1):
                    #nrows1 = temp[0]
                #else:
                nrows1=sheet1.nrows
                ncols1=sheet1.ncols
                #print  str(column1).decode("unicode_escape").encode("utf8")
                #print  str(row1).decode("unicode_escape").encode("utf8")
            else:
                sheet1=workBook1[nameTemp]
                #nrows1 = sheet1.max_row
                #ncols1 = sheet1.max_column
                #temp = SheetSpace(sheet1, flagOld)
                #if (temp[0] != -1):
                    #nrows1 = temp[0]
                #else:
                nrows1 = sheet1.max_row
                #if (temp[1] != -1):
                    #ncols1 = temp[1]
                #else:
                ncols1 = sheet1.max_column
                if nrows1==1 and ncols1==1 and unicode(sheet1.cell(row=1, column=1).value) ==u'None':
                    nrows1=0
                    ncols1=0
                    column1=[]
                    row1=[]
                else:
                    for i in range(1,ncols1+1):
                        Data = sheet1.cell(row=1, column=i).value
                        if type(Data) == type(1L) or type(Data) == type(1):
                            column1.append(unicode(float(Data)))
                        else:
                            column1.append(unicode(Data))
                    for i in range(1, nrows1 + 1):
                        Data = sheet1.cell(row=i, column=1).value
                        if type(Data) == type(1L) or type(Data) == type(1):
                            row1.append(unicode(float(Data)))
                        else:
                            row1.append(unicode(Data))
                #print column1,' ',row1,' ',nrows1,' ',ncols1
            #print nrows1, ncols1
            if flagNew==0:
                sheet2 = workBook2.sheet_by_name(nameTemp)
                #column2 = sheet2.row_values(0)
                #row2 = sheet2.col_values(0)

                try:
                    columntemp2 = sheet2.row_values(0)
                    rowtemp2 = sheet2.col_values(0)
                    for i in columntemp2:
                        column2.append(unicode(i))
                    for i in rowtemp2:
                        row2.append(unicode(i))
                except IndexError:
                    column2 = [u'None']
                    row2 = [u'None']

                nrows2 = sheet2.nrows
                ncols2 = sheet2.ncols
                #temp = SheetSpace(sheet2, flagNew)
                #if (temp[0] != -1):
                    #nrows2 = temp[0]
                #else:
                    #nrows2 = sheet2.nrows
                #if (temp[1] != -1):
                    #ncols2 = temp[1]
                #else:
                    #ncols2 = sheet2.ncols
                #print  str(column2).decode("unicode_escape").encode("utf8")
                #print  str( row2).decode("unicode_escape").encode("utf8")

            else:
                sheet2 = workBook2[nameTemp]
                #temp = SheetSpace(sheet2, flagNew)
                #if (temp[0] != -1):
                    #nrows2 = temp[0]
                #else:
                    #nrows2 = sheet2.max_row
                #if (temp[1] != -1):
                    #ncols2 = temp[1]
                #else:
                    #ncols2 = sheet2.max_column

                nrows2 = sheet2.max_row
                ncols2 = sheet2.max_column

                if nrows2 == 1 and ncols2 == 1 and unicode(sheet2.cell(row=1, column=1).value) == u'None':
                    nrows2 = 0
                    ncols2 = 0
                    column2=[]
                    row2=[]
                else:
                    for i in range(1, ncols2 + 1):
                        Data = sheet2.cell(row=1, column=i).value
                        if type(Data) == type(1L) or type(Data) == type(1):
                            column2.append(unicode(float(Data)))
                        else:
                            column2.append(unicode(Data))
                    for i in range(1, nrows2 + 1):
                        Data = sheet2.cell(row=i, column=1).value
                        if type(Data) == type(1L) or type(Data) == type(1):
                            row2.append(unicode(float(Data)))
                        else:
                            row2.append(unicode(Data))
                #print column2, ' ', row2, ' ', nrows2, ' ', ncols2
            #except xlrd.biffh.XLRDError:
                #names = data.sheet_names()
                #nameTemp=str(nameTemp.encode('utf-8'))
                #print type(nameTemp)
                #sheet1 = workBook1.sheet_by_name(nameTemp)
                #sheet2 = workBook2.sheet_by_name(nameTemp)
            # 默认列表头唯一，以列表头为参考，判断列增删
            #print nrows2, ncols2
            #print column1
            #print column2
            #print row1
            #print row2
            intersection=list(set(column1).intersection(set(column2)))
            #print str(intersection).decode("unicode_escape").encode("utf8")
            # 增列
            columnAdd=list(set(column2).difference(set(intersection)))
            #print str(columnAdd).decode("unicode_escape").encode("utf8")
            # 删列
            columnDel=list(set(column1).difference(set(intersection)))
            #print str(columnDel).decode("unicode_escape").encode("utf8")
            # 默认行表头唯一，以行表头为参考，判断行增删
            intersection = list(set(row1).intersection(set(row2)))
            # 增行
            rowAdd = list(set(row2).difference(set(intersection)))
            #print str(rowAdd).decode("unicode_escape").encode("utf8")
            # 删列行
            rowDel = list(set(row1).difference(set(intersection)))
            #print str(rowDel).decode("unicode_escape").encode("utf8")
            # 定义行列增删有关的字典
            global dicRowAdd,dicColAdd,dicRowDel,dicColDel
            dicRowAdd={}
            dicColAdd={}
            dicRowDel={}
            dicColDel={}
            for i in columnAdd:
                dicColAdd[column2.index(i)]=unicode(i)
            for i in columnDel:
                dicColDel[column1.index(i)]=unicode(i)
            for i in rowAdd:
                dicRowAdd[row2.index(i)]=unicode(i)
            for i in rowDel:
                dicRowDel[row1.index(i)] =unicode(i)
            #print str(dicRowAdd).decode("unicode_escape").encode("utf8")
            #print str(dicColAdd).decode("unicode_escape").encode("utf8")
            #print str(dicRowDel).decode("unicode_escape").encode("utf8")
            #print str(dicColDel).decode("unicode_escape").encode("utf8")
            # 对比表格，需要定义为全局变量，不然会在函数执行完就被销毁
            #global excelWindow1
            #excelWindow1 = MyDialog1()
            #excelWindow1.show()
            #global excelWindow2
            #excelWindow2 = MyDialog2()
            global excelWindow
            excelWindow = QtGui.QMainWindow()
            excelMain = Ui_ExcelMainWindow1()
            excelMain.setupUi(excelWindow)
            excelWindow.show()

        except (AttributeError,xlrd.biffh.XLRDError,IndexError):
            pass

    def aboutLittleDiff(self):
        aboutWindow.show()

    def showToolBar(self):
        global showTool
        if showTool==1:
            self.toolBar.setVisible(False)
            showTool = 0
        else:
            self.toolBar.setVisible(True)
            showTool = 1

    def printExcel(self):
        self.listWidget_Add_Col.clear()
        self.listWidget_Del_Col_2.clear()
        self.listWidget__Add_Row.clear()
        self.listWidget_Del_Row.clear()
        self.listWidget_Del_Cell.clear()
        self.listWidget__Add_Cell.clear()
        __sortingEnabled = self.listWidget_Add_Col.isSortingEnabled()
        self.listWidget_Add_Col.setSortingEnabled(False)
        self.listWidget_Add_Col.setSortingEnabled(__sortingEnabled)
        i=0
        for key in dicColAdd:
            item_Add_Col = QtGui.QListWidgetItem()
            self.listWidget_Add_Col.addItem(item_Add_Col)
            item_Add_Col = self.listWidget_Add_Col.item(i)
            #if type(dicColAdd[key])==u'中文':
            item_Add_Col.setText(_translate("MainWindow", FloatAndUnicode(dicColAdd[key]), None))
            #else:
                #item_Add_Col.setText(_translate("MainWindow", str(dicColAdd[key]), None))
            i=i+1

        __sortingEnabled = self.listWidget_Del_Col_2.isSortingEnabled()
        self.listWidget_Del_Col_2.setSortingEnabled(False)
        self.listWidget_Del_Col_2.setSortingEnabled(__sortingEnabled)
        i = 0
        for key in dicColDel:
            item_Del_Col_2 = QtGui.QListWidgetItem()
            self.listWidget_Del_Col_2.addItem(item_Del_Col_2)
            item_Del_Col_2 = self.listWidget_Del_Col_2.item(i)
            #if type(dicColDel[key]) == u'中文':
            item_Del_Col_2.setText(_translate("MainWindow", FloatAndUnicode(dicColDel[key]), None))
            #else:
                #item_Del_Col_2.setText(_translate("MainWindow", str(dicColDel[key]), None))
            i = i + 1

        __sortingEnabled = self.listWidget__Add_Row.isSortingEnabled()
        self.listWidget__Add_Row.setSortingEnabled(False)
        self.listWidget__Add_Row.setSortingEnabled(__sortingEnabled)
        i = 0
        for key in dicRowAdd:
            item__Add_Row = QtGui.QListWidgetItem()
            self.listWidget__Add_Row.addItem(item__Add_Row)
            item__Add_Row = self.listWidget__Add_Row.item(i)
            #if type(dicRowAdd[key]) == u'中文':
            item__Add_Row.setText(_translate("MainWindow", FloatAndUnicode(dicRowAdd[key]), None))
            #else:
                #item__Add_Row.setText(_translate("MainWindow", str(dicRowAdd[key]), None))
            i = i + 1

        __sortingEnabled = self.listWidget_Del_Cell.isSortingEnabled()
        self.listWidget_Del_Cell.setSortingEnabled(False)
        self.listWidget_Del_Cell.setSortingEnabled(__sortingEnabled)
        i = 0
        for key in dicCellDel:
            item_Del_Cell = QtGui.QListWidgetItem()
            self.listWidget_Del_Cell.addItem(item_Del_Cell)
            item_Del_Cell = self.listWidget_Del_Cell.item(i)
            item_Del_Cell.setText(_translate("MainWindow", FloatAndUnicode(dicCellDel[key])+'--->'+FloatAndUnicode(dicCellAdd[key]), None))
            i = i + 1

        __sortingEnabled = self.listWidget__Add_Cell.isSortingEnabled()
        self.listWidget__Add_Cell.setSortingEnabled(False)
        self.listWidget__Add_Cell.setSortingEnabled(__sortingEnabled)
        i = 0
        for key in dicCellAdd:
            item__Add_Cell = QtGui.QListWidgetItem()
            self.listWidget__Add_Cell.addItem(item__Add_Cell)
            item__Add_Cell = self.listWidget__Add_Cell.item(i)
            item__Add_Cell.setText(_translate("MainWindow", FloatAndUnicode(dicCellDel[key])+'--->'+FloatAndUnicode(dicCellAdd[key]), None))
            i = i + 1

        __sortingEnabled = self.listWidget_Del_Row.isSortingEnabled()
        self.listWidget_Del_Row.setSortingEnabled(False)
        self.listWidget_Del_Row.setSortingEnabled(__sortingEnabled)
        i = 0
        for key in dicRowDel:
            item_Del_Row = QtGui.QListWidgetItem()
            self.listWidget_Del_Row.addItem(item_Del_Row)
            item_Del_Row = self.listWidget_Del_Row.item(i)
            #if type(dicRowDel[key]) == u'中文':
            item_Del_Row.setText(_translate("MainWindow", FloatAndUnicode(dicRowDel[key]), None))
            #else:
                #item_Del_Row.setText(_translate("MainWindow", str(dicRowDel[key]), None))
            i = i + 1

    def saveExcel(self):
        global dicRowAdd, dicColAdd, dicRowDel, dicColDel, dicCellAdd, dicCellDel
        with open('Revise.log', 'w') as f:  # 如果filename不存在会自动创建
            f.write('新增行号：\n')
            for key in dicRowAdd:
                f.write(str(key))
                f.write(' ')
            f.write('\n')
            f.write('删除行号：\n')
            for key in dicRowDel:
                f.write(str(key))
                f.write(' ')
            f.write('\n')
            f.write('新增列号：\n')
            for key in dicColAdd:
                f.write(str(key))
                f.write(' ')
            f.write('\n')
            f.write('删除列号：\n')
            for key in dicColDel:
                f.write(str(key))
                f.write(' ')
            f.write('\n')
    # 没能实现打印机功能调用
    '''
    def printLog(self):
        file = QtCore.QFile("Revise.log")
        if file.open(QtCore.QIODevice.ReadOnly | QtCore.QIODevice.Text):
            textStream1 = QtCore.QTextStream(file)
        while not textStream1.atEnd():
            self.cetext.append(textStream1.readLine())
        file.close()

        printer = QtGui.QPrinter()
        printDialog = QtGui.QPrintDialog(printer, self)
        if printDialog.exec_():
            doc = self.text.document()
            doc.print_(printer)
    '''

    def showExcel(self):
        global dicRowAdd, dicColAdd, dicRowDel, dicColDel, dicCellAdd, dicCellDel, nrows1, ncols1, nrows2, ncols2
        global excelWindow
        temRows=nrows1+len(dicRowAdd)#列
        temCols=ncols1+len(dicColAdd)
        # 定义一个sender可以判断信号是从哪里发射的，这样多个信号可以绑定一个槽函数
        sender = QtGui.QMainWindow().sender()
        index = sender.currentRow()
        # 标识 被删除行
        if sender==self.listWidget_Del_Row:
            value=dicRowDel.values()[index]
            for i,j in dicRowDel1.items():
                if j==value:
                    temp=i
                    break
            ran=(temp,0,temp,temCols-1)
            excelWindow.close()
            excelMain = Ui_ExcelMainWindow1()
            excelMain.setupUi(excelWindow)
            excelMain.select3(ran)

            excelWindow.show()
            excelMain.select3(ran)
        # 标识 增加列
        elif(sender==self.listWidget_Add_Col):
            value = dicColAdd.values()[index]
            for i, j in dicColAdd1.items():
                if j == value:
                    temp = i
                    break
            ran = (0, temp, temRows-1, temp)
            excelWindow.close()
            excelMain = Ui_ExcelMainWindow1()
            excelMain.setupUi(excelWindow)
            excelMain.select3(ran)

            excelWindow.show()
            excelMain.select3(ran)
        # 标识 增加列
        elif(sender==self.listWidget__Add_Row):
            value = dicRowAdd.values()[index]
            for i, j in dicRowAdd1.items():
                if j == value:
                    temp = i
                    break
            ran = (temp, 0, temp, temCols - 1)
            excelWindow.close()
            excelMain = Ui_ExcelMainWindow1()
            excelMain.setupUi(excelWindow)
            excelMain.select3(ran)

            excelWindow.show()
            excelMain.select3(ran)

        # 标识 删除行
        elif(sender==self.listWidget_Del_Col_2):
            value = dicColDel.values()[index]
            for i, j in dicColDel1.items():
                if j == value:
                    temp = i
                    break
            ran = (0, temp, temRows - 1, temp)
            excelWindow.close()
            excelMain = Ui_ExcelMainWindow1()
            excelMain.setupUi(excelWindow)
            excelMain.select3(ran)

            excelWindow.show()
            excelMain.select3(ran)

        # 标识 表1中的单元格
        elif(sender==self.listWidget_Del_Cell):
            value = dicCellDel.values()[index]
            for i, j in dicCellDel.items():
                if j == value:
                    temp = i
                    break
            ran = (temp[0], temp[1], temp[0], temp[1])
            excelWindow.close()
            excelMain = Ui_ExcelMainWindow1()

            excelMain.setupUi(excelWindow)
            excelMain.select3(ran)
            excelWindow.show()
            excelMain.select3(ran)

        # 标识 表2中的单元格
        else:
            value = dicCellAdd.values()[index]
            for i, j in dicCellAdd.items():
                if j == value:
                    temp = i
                    break
            ran = (temp[0], temp[1], temp[0], temp[1])
            excelWindow.close()
            excelMain = Ui_ExcelMainWindow1()

            excelMain.setupUi(excelWindow)
            excelMain.select3(ran)
            excelWindow.show()
            excelMain.select3(ran)


class Ui_ExcelMainWindow1(object):
    def setupUi(self, ExcelMainWindow):
        ExcelMainWindow.setObjectName(_fromUtf8("ExcelMainWindow"))
        #ExcelMainWindow.resize(771, 729)
        #self.setGeometry(QtCore.QRect(1300, 100, 771, 729))
        desktop = QtGui.QApplication.desktop()
        # print desktop.height(),desktop.width()
        # 尽量适配屏幕
        ExcelMainWindow.setGeometry(QtCore.QRect(desktop.width() / 2 + 300, desktop.height()/2-300, 771, 729))
        ExcelMainWindow.setWindowIcon(QtGui.QIcon(':/ico/Cut.ico'))
        ExcelMainWindow.setStyleSheet(_fromUtf8("background-color: rgb(255, 255, 255);"))
        self.centralwidget = QtGui.QWidget(ExcelMainWindow)
        self.centralwidget.setObjectName(_fromUtf8("centralwidget"))

        global dicRowAdd, dicColAdd, dicRowDel, dicColDel, dicCellAdd, dicCellDel, nrows1, ncols1, nrows2, ncols2
        dicCellDel = {}
        # dicCellAdd={}
        # 记录增删数据在对比表格中的位置
        global dicRowAdd1, dicColAdd1, dicRowDel1, dicColDel1



        # 只需遍历一次这个表格，若表格大小为x*y，则算法时间复杂度为O(m*n)
        # ---------------------------------------------------------------------------------------------------#
        # 整个算法过程可以描述为：从(1,A)单元格处开始，以每行为单位，从左向右输出结果。在输出每行结果时，首先判断 #
        # 该行是否进行过增删操作。若没有则比较旧表与新表的该行内容，记录发生过变动的单元格；若发生过增删操作，则分 #
        # 两种情况讨论。逐行输出的过程中需要记录绘制“增加行”，“删除行”，“增加单元格”，“删除单元格”的次数， #
        # 以保证比较的行与单元格在两个表中是的相对应的；生成表格过程中会记录被修改过的行，列，单元格在结果中的实际 #
        # 位置（原位置+偏移）                                                                                 #
        # ---------------------------------------------------------------------------------------------------#

        global sheet1, sheet2, flagOld, flagNew
        #存在同名表，有一张为空的情况
        if nrows1==0 and ncols1==0 and flagNew==0:
            self.tableWidget = QtGui.QTableWidget(self.centralwidget)
            self.tableWidget.setGeometry(QtCore.QRect(20, 40, 731, 321))
            self.tableWidget.setObjectName(_fromUtf8("tableWidget"))
            self.tableWidget.setColumnCount(ncols2 )
            self.tableWidget.setRowCount(nrows2 )
            self.tableWidget.verticalHeader().setClickable(True)
            self.tableWidget.horizontalHeader().setClickable(True)
            font = self.tableWidget.horizontalHeader().font()
            font.setBold(True)
            self.tableWidget.horizontalHeader().setFont(font)
            self.tableWidget.verticalHeader().setFont(font)
            self.tableWidget.clear()

            self.tableWidget_2 = QtGui.QTableWidget(self.centralwidget)
            self.tableWidget_2.setGeometry(QtCore.QRect(20, 400, 731, 321))
            self.tableWidget_2.setObjectName(_fromUtf8("tableWidget_2"))
            self.tableWidget_2.setColumnCount(ncols2 )
            # print ncols1 + len(dicColAdd),nrows1+len(dicRowAdd)
            # print ncols1,nrows1,len(dicColAdd),len(dicRowAdd)
            self.tableWidget_2.setRowCount(nrows2 )
            self.tableWidget_2.verticalHeader().setClickable(True)
            self.tableWidget_2.horizontalHeader().setClickable(True)
            font = self.tableWidget_2.horizontalHeader().font()
            font.setBold(True)
            self.tableWidget_2.horizontalHeader().setFont(font)
            self.tableWidget_2.verticalHeader().setFont(font)
            self.tableWidget_2.clear()
            #print 1
            labelCol=[]
            labelRow=[]
            for i in range(0,nrows2):
                labelRow.append(' ')
            for i in range(0,ncols2):
                labelCol.append(' ')
            self.tableWidget.setHorizontalHeaderLabels(labelCol)
            self.tableWidget.setVerticalHeaderLabels(labelRow)
            for i in range(0,nrows2):
                temp=sheet2.row_values(i)
                num=0
                for j in temp:
                    if type(j) == type(1L) or type(j) == type(1):
                        newItem = QtGui.QTableWidgetItem(unicode(float(j)))
                    else:
                        newItem = QtGui.QTableWidgetItem(unicode(j))

                    newItem.setBackgroundColor(QtGui.QColor(0, 191, 255))

                    newItem2 = QtGui.QTableWidgetItem((' '))
                    newItem2.setBackgroundColor(QtGui.QColor(255, 20, 147))
                    self.tableWidget_2.setItem(i, num, newItem)
                    self.tableWidget.setItem(i, num, newItem2)
                    num=num+1

        elif nrows2 == 0 and ncols2 == 0 and flagOld == 0:
            self.tableWidget = QtGui.QTableWidget(self.centralwidget)
            self.tableWidget.setGeometry(QtCore.QRect(20, 40, 731, 321))
            self.tableWidget.setObjectName(_fromUtf8("tableWidget"))
            self.tableWidget.setColumnCount(ncols1 )
            self.tableWidget.setRowCount(nrows1)
            self.tableWidget.verticalHeader().setClickable(True)
            self.tableWidget.horizontalHeader().setClickable(True)
            font = self.tableWidget.horizontalHeader().font()
            font.setBold(True)
            self.tableWidget.horizontalHeader().setFont(font)
            self.tableWidget.verticalHeader().setFont(font)
            self.tableWidget.clear()

            self.tableWidget_2 = QtGui.QTableWidget(self.centralwidget)
            self.tableWidget_2.setGeometry(QtCore.QRect(20, 400, 731, 321))
            self.tableWidget_2.setObjectName(_fromUtf8("tableWidget_2"))
            self.tableWidget_2.setColumnCount(ncols1 )
            # print ncols1 + len(dicColAdd),nrows1+len(dicRowAdd)
            # print ncols1,nrows1,len(dicColAdd),len(dicRowAdd)
            self.tableWidget_2.setRowCount(nrows1 )
            self.tableWidget_2.verticalHeader().setClickable(True)
            self.tableWidget_2.horizontalHeader().setClickable(True)
            font = self.tableWidget_2.horizontalHeader().font()
            font.setBold(True)
            self.tableWidget_2.horizontalHeader().setFont(font)
            self.tableWidget_2.verticalHeader().setFont(font)
            self.tableWidget_2.clear()
            #print 2
            labelCol = []
            labelRow = []
            for i in range(0, nrows1):
                labelRow.append(' ')
            for i in range(0, ncols1):
                labelCol.append(' ')
            self.tableWidget_2.setHorizontalHeaderLabels(labelCol)
            self.tableWidget_2.setVerticalHeaderLabels(labelRow)
            for i in range(0, nrows1):
                temp = sheet1.row_values(i)
                num = 0
                #print temp
                for j in temp:
                    #print 'lie',j
                    if type(j) == type(1L) or type(j) == type(1):
                        newItem = QtGui.QTableWidgetItem(unicode(float(j)))
                    else:
                        newItem = QtGui.QTableWidgetItem(unicode(j))

                    newItem.setBackgroundColor(QtGui.QColor(255, 20, 147))
                    self.tableWidget.setItem(i, num, newItem)

                    newItem2 = QtGui.QTableWidgetItem((' '))
                    newItem2.setBackgroundColor(QtGui.QColor(0, 191, 255))
                    self.tableWidget_2.setItem(i, num, newItem2)
                    #print 'clear'
                    num = num + 1
                    #print num
        elif nrows1 == 0 and ncols1 == 0 and flagNew == 1:
            self.tableWidget = QtGui.QTableWidget(self.centralwidget)
            self.tableWidget.setGeometry(QtCore.QRect(20, 40, 731, 321))
            self.tableWidget.setObjectName(_fromUtf8("tableWidget"))
            self.tableWidget.setColumnCount(ncols2)
            self.tableWidget.setRowCount(nrows2)
            self.tableWidget.verticalHeader().setClickable(True)
            self.tableWidget.horizontalHeader().setClickable(True)
            font = self.tableWidget.horizontalHeader().font()
            font.setBold(True)
            self.tableWidget.horizontalHeader().setFont(font)
            self.tableWidget.verticalHeader().setFont(font)
            self.tableWidget.clear()

            self.tableWidget_2 = QtGui.QTableWidget(self.centralwidget)
            self.tableWidget_2.setGeometry(QtCore.QRect(20, 400, 731, 321))
            self.tableWidget_2.setObjectName(_fromUtf8("tableWidget_2"))
            self.tableWidget_2.setColumnCount(ncols2)
            # print ncols1 + len(dicColAdd),nrows1+len(dicRowAdd)
            # print ncols1,nrows1,len(dicColAdd),len(dicRowAdd)
            self.tableWidget_2.setRowCount(nrows2)
            self.tableWidget_2.verticalHeader().setClickable(True)
            self.tableWidget_2.horizontalHeader().setClickable(True)
            font = self.tableWidget_2.horizontalHeader().font()
            font.setBold(True)
            self.tableWidget_2.horizontalHeader().setFont(font)
            self.tableWidget_2.verticalHeader().setFont(font)
            self.tableWidget_2.clear()
            #print 3
            labelCol = []
            labelRow = []
            for i in range(0, nrows2):
                labelRow.append(' ')
            for i in range(0, ncols2):
                labelCol.append(' ')
            self.tableWidget.setHorizontalHeaderLabels(labelCol)
            self.tableWidget.setVerticalHeaderLabels(labelRow)
            for i in range(0, nrows2):
                #print 'i',i
                for j in range(0,ncols2):
                    #print 'j', j
                    Data = sheet2.cell(row=i+1, column=j+1).value
                    #print Data
                    if type(Data) == type(1L) or type(Data) == type(1):
                        newItem = QtGui.QTableWidgetItem(unicode(float(Data)))
                    else:
                        newItem = QtGui.QTableWidgetItem(unicode(Data))

                    newItem.setBackgroundColor(QtGui.QColor(0, 191, 255))


                    newItem2 = QtGui.QTableWidgetItem((' '))
                    newItem2.setBackgroundColor(QtGui.QColor(255, 20, 147))
                    self.tableWidget_2.setItem(i,j, newItem)
                    self.tableWidget.setItem(i, j, newItem2)
                    #print 'clear'
            #print 'suss'
        elif nrows2 == 0 and ncols2 == 0 and flagOld == 1:
            self.tableWidget = QtGui.QTableWidget(self.centralwidget)
            self.tableWidget.setGeometry(QtCore.QRect(20, 40, 731, 321))
            self.tableWidget.setObjectName(_fromUtf8("tableWidget"))
            self.tableWidget.setColumnCount(ncols1 )
            self.tableWidget.setRowCount(nrows1 )
            self.tableWidget.verticalHeader().setClickable(True)
            self.tableWidget.horizontalHeader().setClickable(True)
            font = self.tableWidget.horizontalHeader().font()
            font.setBold(True)
            self.tableWidget.horizontalHeader().setFont(font)
            self.tableWidget.verticalHeader().setFont(font)
            self.tableWidget.clear()

            self.tableWidget_2 = QtGui.QTableWidget(self.centralwidget)
            self.tableWidget_2.setGeometry(QtCore.QRect(20, 400, 731, 321))
            self.tableWidget_2.setObjectName(_fromUtf8("tableWidget_2"))
            self.tableWidget_2.setColumnCount(ncols1 )
            # print ncols1 + len(dicColAdd),nrows1+len(dicRowAdd)
            # print ncols1,nrows1,len(dicColAdd),len(dicRowAdd)
            self.tableWidget_2.setRowCount(nrows1 )
            self.tableWidget_2.verticalHeader().setClickable(True)
            self.tableWidget_2.horizontalHeader().setClickable(True)
            font = self.tableWidget_2.horizontalHeader().font()
            font.setBold(True)
            self.tableWidget_2.horizontalHeader().setFont(font)
            self.tableWidget_2.verticalHeader().setFont(font)
            self.tableWidget_2.clear()
            #print nrows1,ncols1
            labelCol = []
            labelRow = []
            for i in range(0, nrows1):
                labelRow.append(' ')
            for i in range(0, ncols1):
                labelCol.append(' ')
            self.tableWidget_2.setHorizontalHeaderLabels(labelCol)
            self.tableWidget_2.setVerticalHeaderLabels(labelRow)
            for i in range(0, nrows1):
                # print 'i',i
                for j in range(0, ncols1):
                    # print 'j', j
                    Data = sheet1.cell(row=i + 1, column=j + 1).value
                    # print Data
                    if type(Data) == type(1L) or type(Data) == type(1):
                        newItem = QtGui.QTableWidgetItem(unicode(float(Data)))
                    else:
                        newItem = QtGui.QTableWidgetItem(unicode(Data))

                    newItem.setBackgroundColor(QtGui.QColor(255, 20, 147))
                    self.tableWidget.setItem(i, j, newItem)

                    newItem2 = QtGui.QTableWidgetItem((' '))
                    newItem2.setBackgroundColor(QtGui.QColor(0, 191, 255))
                    self.tableWidget_2.setItem(i, j, newItem2)
        else:
            #####################
            self.tableWidget = QtGui.QTableWidget(self.centralwidget)
            self.tableWidget.setGeometry(QtCore.QRect(20, 40, 731, 321))
            self.tableWidget.setObjectName(_fromUtf8("tableWidget"))
            self.tableWidget.setColumnCount(ncols1 + len(dicColAdd))
            self.tableWidget.setRowCount(nrows1 + len(dicRowAdd))
            self.tableWidget.verticalHeader().setClickable(True)
            self.tableWidget.horizontalHeader().setClickable(True)
            font = self.tableWidget.horizontalHeader().font()
            font.setBold(True)
            self.tableWidget.horizontalHeader().setFont(font)
            self.tableWidget.verticalHeader().setFont(font)
            self.tableWidget.clear()

            self.tableWidget_2 = QtGui.QTableWidget(self.centralwidget)
            self.tableWidget_2.setGeometry(QtCore.QRect(20, 400, 731, 321))
            self.tableWidget_2.setObjectName(_fromUtf8("tableWidget_2"))
            self.tableWidget_2.setColumnCount(ncols1 + len(dicColAdd))
            # print ncols1 + len(dicColAdd),nrows1+len(dicRowAdd)
            # print ncols1,nrows1,len(dicColAdd),len(dicRowAdd)
            self.tableWidget_2.setRowCount(nrows1 + len(dicRowAdd))
            self.tableWidget_2.verticalHeader().setClickable(True)
            self.tableWidget_2.horizontalHeader().setClickable(True)
            font = self.tableWidget_2.horizontalHeader().font()
            font.setBold(True)
            self.tableWidget_2.horizontalHeader().setFont(font)
            self.tableWidget_2.verticalHeader().setFont(font)
            self.tableWidget_2.clear()
            labelRow = []
            labelCol = []
            k1 = 0  # 记录绘制结果过程中，绘制“增加行”的次数
            k2 = 0  # 记录绘制结果过程中，绘制“删除行”的次数
            #k1max=len(dicRowAdd)
            #k2max=len(dicRowDel)
            #print k1max,k2max
            #print nrows1 , nrows2
            for i in range(0, nrows1 + len(dicRowAdd)):
                # 记录列增，删次数
                m = 0  # 记录绘制结果过程中，绘制“增加单元格”的次数
                n = 0  # 记录绘制结果过程中，绘制“删除单元格”的次数
                #temp1 = sheet1.row_values(i-k1)
                #temp2 = sheet2.row_values(i-k2)
                #print 'i-k1',i-k1
                #print 'i-k2',i-k2
                temp1 = GetSheetRowVal(sheet1, i - k1, flagOld, ncols1,nrows1)
                temp2 = GetSheetRowVal(sheet2, i - k2, flagNew, ncols2,nrows2)
                #print str(temp1).decode("unicode_escape").encode("utf8")
                #print str(temp2).decode("unicode_escape").encode("utf8")
                #print ' '
                #print '   ',ncols1 + len(dicColAdd)
                for j in range(0, ncols1 + len(dicColAdd)):
                    # 判断行列增删
                    #if i==nrows1 + len(dicRowAdd) and temp1[0]!=temp2[0]:

                    if dicRowAdd.has_key(i - k2) and temp2!=[]:
                        #print'1'
                        # 第一次进入循环时，输出“增加行”次数加一
                        if j == 0:
                            #if k1<k1max and i!=nrows1 + len(dicRowAdd):
                            k1 = k1 + 1
                                #print 'k1:',k1
                            value = dicRowAdd[i - k2]
                            dicRowAdd1[i] = value
                            labelRow.append(' ')
                        # 输出蓝色的空行
                        if dicColDel.has_key(j - m):
                            #print'4'
                            # 遇到删除的列,输出带红色背景的数据,并计数
                            if temp1!=[]:
                                newItem = QtGui.QTableWidgetItem(temp1[j - m])
                            else:
                                newItem = QtGui.QTableWidgetItem(temp2[j - m])
                            newItem.setBackgroundColor(QtGui.QColor(255, 20, 147))
                            self.tableWidget.setItem(i, j, newItem)
                            n = n + 1
                            if i == 0:
                                value = dicColDel[j - m]
                                dicColDel1[j] = value
                                labelCol.append(str(j - m))
                        elif dicColAdd.has_key(j - n):
                            #print'5'
                            # 遇到增加的列，仍然输出蓝色，并计数
                            newItem = QtGui.QTableWidgetItem(' ')
                            newItem.setBackgroundColor(QtGui.QColor(0, 191, 255))
                            self.tableWidget.setItem(i, j, newItem)
                            m = m + 1
                            if i == 0:
                                value = dicColAdd[j - n]
                                dicColAdd1[j] = value
                                labelCol.append(' ')
                            #if i==nrows1 + len(dicRowAdd) and temp1[0]!=temp2[0]:
                                #newItem = QtGui.QTableWidgetItem(temp1[j - m])
                                #newItem.setBackgroundColor(QtGui.QColor(255, 20, 147))
                                #self.tableWidget.setItem(i+1, j, newItem)
                        else:
                            #print'6'
                            newItem = QtGui.QTableWidgetItem(' ')
                            newItem.setBackgroundColor(QtGui.QColor(0, 191, 255))
                            self.tableWidget.setItem(i, j, newItem)
                            if i == 0:
                                labelCol.append(str(j - m))

                    elif dicRowDel.has_key(i - k1)and temp1!=[]:
                       # print'2'
                        if j == 0:
                            #if k2 < k2max and i!=nrows1 + len(dicRowAdd):
                            k2 = k2 + 1
                            #k2 = k2 + 1
                            #print 'k2:', k2
                            value = dicRowDel[i - k1]
                            dicRowDel1[i] = value
                            labelRow.append(str(i - k1))
                        # 输出红色的空行
                        if dicColDel.has_key(j - n):
                            #print '7'
                            # 遇到删除的列，仍然输出红色，并计数
                            newItem = QtGui.QTableWidgetItem(temp1[j - n])
                            newItem.setBackgroundColor(QtGui.QColor(255, 20, 147))
                            self.tableWidget.setItem(i, j, newItem)
                            m = m + 1
                            if i == 0:
                                value = dicColDel[j - n]
                                dicColDel1[j] = value
                                labelCol.append(str(j - n))
                        elif dicColAdd.has_key(j - m):
                            #print '8'
                            # 遇到增加的列,输出蓝色，并计数
                            newItem = QtGui.QTableWidgetItem(' ')
                            newItem.setBackgroundColor(QtGui.QColor(0, 191, 255))
                            self.tableWidget.setItem(i, j, newItem)
                            n = n + 1
                            if i == 0:
                                value = dicColAdd[j - m]
                                dicColAdd1[j] = value
                                labelCol.append(' ')

                        else:
                            #print '9'
                            newItem = QtGui.QTableWidgetItem(temp1[j - n])
                            newItem.setBackgroundColor(QtGui.QColor(255, 20, 147))
                            self.tableWidget.setItem(i, j, newItem)
                            if i == 0:
                                labelCol.append(str(j - n))
                    else:
                        # 正常输出 比较出修改的cell
                        #print'3'
                        if j == 0:
                            labelRow.append(str(i - k1))
                        if dicColDel.has_key(j - n):
                            #print '10'
                            # 遇到删除的列，输出红色，并计数
                            newItem = QtGui.QTableWidgetItem(temp1[j - n])
                            newItem.setBackgroundColor(QtGui.QColor(255, 20, 147))
                            self.tableWidget.setItem(i, j, newItem)
                            m = m + 1
                            if i == 0:
                                value = dicColDel[j - n]
                                dicColDel1[j] = value
                                # labelRow.append(i - k1)
                                labelCol.append(str(j - n))
                        elif dicColAdd.has_key(j - m):
                            #print '11'
                            # 遇到增加的列,输出蓝色，并计数
                            newItem = QtGui.QTableWidgetItem(' ')
                            newItem.setBackgroundColor(QtGui.QColor(0, 191, 255))
                            self.tableWidget.setItem(i, j, newItem)
                            n = n + 1
                            if i == 0:
                                value = dicColAdd[j - m]
                                dicColAdd1[j] = value
                                labelCol.append(' ')
                        else:
                            # '12'
                            if i == 0:
                                labelCol.append(str(j - n))
                            if ((temp1[j - n])) == ((temp2[j - m])):
                                newItem = QtGui.QTableWidgetItem(temp1[j - n])
                                self.tableWidget.setItem(i, j, newItem)
                            else:
                                newItem = QtGui.QTableWidgetItem(temp1[j - n])
                                newItem.setBackgroundColor(QtGui.QColor(255, 255, 0))
                                self.tableWidget.setItem(i, j, newItem)
                                # 记录发生变动的单元格
                                dicCellDel[(i, j)] = temp1[j - n]
            #layout = QtGui.QHBoxLayout()
            #layout.addWidget(self.tableWidget)
            #self.setLayout(layout)
            # print labelCol
            # print labelRow
            self.tableWidget.setHorizontalHeaderLabels(labelCol)
            self.tableWidget.setVerticalHeaderLabels(labelRow)
            #print 'over1'
            # dicCellDel = {}
            dicCellAdd = {}
            '''''
            self.tableWidget_2 = QtGui.QTableWidget(self.centralwidget)
            self.tableWidget_2.setGeometry(QtCore.QRect(20, 400, 731, 321))
            self.tableWidget_2.setObjectName(_fromUtf8("tableWidget_2"))
            self.tableWidget_2.setColumnCount(ncols1+len(dicColAdd))
            self.tableWidget_2.setRowCount(nrows1+len(dicRowAdd))

            self.tableWidget_2.verticalHeader().setClickable(True)
            self.tableWidget_2.horizontalHeader().setClickable(True)
            font = self.tableWidget_2.horizontalHeader().font()
            font.setBold(True)
            self.tableWidget_2.horizontalHeader().setFont(font)
            self.tableWidget_2.verticalHeader().setFont(font)
            self.tableWidget_2.clear()
            '''''
            labelRow = []
            labelCol = []
            k1 = 0  # 记录绘制结果过程中，绘制“增加行”的次数
            k2 = 0  # 记录绘制结果过程中，绘制“删除行”的次数
            #k1max=len(dicRowAdd)
            #k2max=len(dicRowDel)
            for i in range(0, nrows1 + len(dicRowAdd)):
                m = 0  # 记录绘制结果过程中，绘制“增加单元格”的次数
                n = 0  # 记录绘制结果过程中，绘制“删除单元格”的次数
                # temp1 = sheet1.row_values(i-k1)
                # temp2 = sheet2.row_values(i-k2)
                temp1 = GetSheetRowVal(sheet1, i - k1, flagOld, ncols1,nrows1)
                temp2 = GetSheetRowVal(sheet2, i - k2, flagNew, ncols2,nrows2)
                #print str(temp1).decode("unicode_escape").encode("utf8")
                #print str(temp2).decode("unicode_escape").encode("utf8")
                #print ' ',i,nrows1, nrows2
                for j in range(0, ncols1 + len(dicColAdd)):
                    if dicRowAdd.has_key(i - k2)and temp2!=[]:
                        if j == 0:
                            #if k1<k1max and i!=nrows1 + len(dicRowAdd):
                            k1 = k1 + 1
                            labelRow.append(str(i - k2))
                        if dicColDel.has_key(j - m):
                            #print '1'
                            newItem = QtGui.QTableWidgetItem(' ')
                            newItem.setBackgroundColor(QtGui.QColor(255, 20, 147))
                            self.tableWidget_2.setItem(i, j, newItem)
                            n = n + 1
                            if i == 0:
                                labelCol.append(' ')
                        elif dicColAdd.has_key(j - n):
                            #print '2'
                            newItem = QtGui.QTableWidgetItem(temp2[j - n])
                            newItem.setBackgroundColor(QtGui.QColor(0, 191, 255))
                            self.tableWidget_2.setItem(i, j, newItem)
                            m = m + 1
                            if i == 0:
                                labelCol.append(str(j - n))
                        else:
                            #print '3'
                            newItem = QtGui.QTableWidgetItem(temp2[j - n])
                            newItem.setBackgroundColor(QtGui.QColor(0, 191, 255))
                            self.tableWidget_2.setItem(i, j, newItem)
                            if i == 0:
                                labelCol.append(str(j - n))
                            #if i == nrows1 + len(dicRowAdd) and temp1[0] != temp2[0]:
                                #newItem = QtGui.QTableWidgetItem(temp1[j - m])
                                #newItem.setBackgroundColor(QtGui.QColor(255, 20, 147))
                                #self.tableWidget.setItem(i + 1, j, newItem)
                    elif dicRowDel.has_key(i - k1)and temp1!=[]:
                        if j == 0:
                            #if k2<k2max and i!=nrows1 + len(dicRowAdd):
                            k2 = k2 + 1
                            labelRow.append(' ')
                        if dicColDel.has_key(j - n):
                            #print '4'
                            newItem = QtGui.QTableWidgetItem(' ')
                            newItem.setBackgroundColor(QtGui.QColor(255, 20, 147))
                            self.tableWidget_2.setItem(i, j, newItem)
                            m = m + 1
                            if i == 0:
                                labelCol.append(' ')
                        elif dicColAdd.has_key(j - m):
                            #print '5'
                            #print str(temp2[j - m]).decode("unicode_escape").encode("utf8")
                            if temp2==[]:
                                newItem = QtGui.QTableWidgetItem(temp1[j - m])
                            else:
                                newItem = QtGui.QTableWidgetItem(temp2[j - m])
                            newItem.setBackgroundColor(QtGui.QColor(0, 191, 255))
                            self.tableWidget_2.setItem(i, j, newItem)
                            n = n + 1
                            #print 'over2'
                            if i == 0:
                                labelCol.append(str(j - m))
                        else:
                            #print '6'
                            newItem = QtGui.QTableWidgetItem(' ')
                            newItem.setBackgroundColor(QtGui.QColor(255, 20, 147))
                            self.tableWidget_2.setItem(i, j, newItem)
                            if i == 0:
                                labelCol.append(' ')
                    else:
                        if j == 0:
                            labelRow.append(str(i - k2))
                        if dicColDel.has_key(j - n):
                            #print '7'
                            newItem = QtGui.QTableWidgetItem(' ')
                            newItem.setBackgroundColor(QtGui.QColor(255, 20, 147))
                            self.tableWidget_2.setItem(i, j, newItem)
                            m = m + 1
                            if i == 0:
                                labelCol.append(' ')
                        elif dicColAdd.has_key(j - m):
                            #print '8'
                            newItem = QtGui.QTableWidgetItem(temp2[j - m])
                            newItem.setBackgroundColor(QtGui.QColor(0, 191, 255))
                            self.tableWidget_2.setItem(i, j, newItem)
                            n = n + 1
                            #print 'over2'
                            if i == 0:
                                labelCol.append(str(j - m))
                        else:
                            #print '9'
                            if i == 0:
                                labelCol.append(str(j - m))
                            if ((temp1[j - n])) == ((temp2[j - m])):

                                newItem = QtGui.QTableWidgetItem(unicode(FloatAndUnicode(temp2[j - m])))
                                self.tableWidget_2.setItem(i, j, newItem)
                            else:
                                newItem = QtGui.QTableWidgetItem((temp2[j - m]))
                                newItem.setBackgroundColor(QtGui.QColor(255, 255, 0))
                                self.tableWidget_2.setItem(i, j, newItem)
                                dicCellAdd[(i, j)] = temp2[j - m]



            #print '1'
            #print str(dicCellDel).decode("unicode_escape").encode("utf8")
            #print str(dicCellAdd).decode("unicode_escape").encode("utf8")

            self.tableWidget_2.setHorizontalHeaderLabels(labelCol)
            self.tableWidget_2.setVerticalHeaderLabels(labelRow)
            #print 'over'
        p1=self.tableWidget_2.verticalScrollBar()
        p2=self.tableWidget.verticalScrollBar()
        h1=self.tableWidget_2.horizontalScrollBar()
        h2=self.tableWidget.horizontalScrollBar()
        m1=self.tableWidget.horizontalHeader()
        m2 = self.tableWidget_2.horizontalHeader()
        n1 = self.tableWidget.verticalHeader()
        n2 = self.tableWidget_2.verticalHeader()

        self.lineEdit = QtGui.QLineEdit(self.centralwidget)
        self.lineEdit.setGeometry(QtCore.QRect(330, 10, 113, 21))
        self.lineEdit.setObjectName(_fromUtf8("lineEdit"))
        self.lineEdit_2 = QtGui.QLineEdit(self.centralwidget)
        self.lineEdit_2.setGeometry(QtCore.QRect(330, 370, 113, 21))
        self.lineEdit_2.setObjectName(_fromUtf8("lineEdit_2"))
        ExcelMainWindow.setCentralWidget(self.centralwidget)

        QtCore.QObject.connect(p1, QtCore.SIGNAL(_fromUtf8("valueChanged(int)")),p2,QtCore.SLOT("setValue(int)"))
        QtCore.QObject.connect(p2, QtCore.SIGNAL(_fromUtf8("valueChanged(int)")),
                               p1, QtCore.SLOT("setValue(int)"))
        QtCore.QObject.connect(h1, QtCore.SIGNAL(_fromUtf8("valueChanged(int)")), h2, QtCore.SLOT("setValue(int)"))
        QtCore.QObject.connect(h2, QtCore.SIGNAL(_fromUtf8("valueChanged(int)")),
                               h1, QtCore.SLOT("setValue(int)"))

        #QtCore.QObject.connect(m1,QtCore.SIGNAL('sectionClicked(int)'),self.moveTo2)
        #QtCore.QObject.connect(m2, QtCore.SIGNAL('sectionClicked(int)'),self.moveTo1)

        self.retranslateUi(ExcelMainWindow)
        QtCore.QMetaObject.connectSlotsByName(ExcelMainWindow)

    def moveTo1(self):
        print '1'
        ran=self.tableWidget.selectedRanges()
        print type(ran)
        #ranges=QtGui.QTableWidgetSelectionRange(ran[0],ran[1],ran[2],ran[3])
        #self.tableWidget_2.setRangeSelected(ranges, True)

    def moveTo2(self):
        print '1'
        ran = self.tableWidget_2.selectedRanges()
        print type(ran)
        #ranges=QtGui.QTableWidgetSelectionRange(ran[0],ran[1],ran[2],ran[3])
        #self.tableWidget.setRangeSelected(ranges, True)

    def retranslateUi(self, ExcelMainWindow):
        ExcelMainWindow.setWindowTitle(_translate("ExcelMainWindow", "Resualt", None))
        self.lineEdit.setText(_translate("ExcelMainWindow", "  Old Table", None))
        self.lineEdit_2.setText(_translate("ExcelMainWindow", "  New Table", None))

    def select1(self, ran):
        global nrows1, dicRowAdd, ncols1, dicColAdd
        #self.tableWidget = QtGui.QTableWidget(self.centralwidget)
        init = QtGui.QTableWidgetSelectionRange(0, 0, nrows1 + len(dicRowAdd) - 1, ncols1 + len(dicColAdd) - 1)
        self.tableWidget.setRangeSelected(init, False)

        ranges = QtGui.QTableWidgetSelectionRange(ran[0], ran[1], ran[2], ran[3])
        self.tableWidget.setRangeSelected(ranges, True)
        self.tableWidget.verticalScrollBar().setSliderPosition(ran[0])
        #self.tableWidget_2.verticalScrollBar().setSliderPosition(ran[0])

    def select2(self, ran):
        global nrows1, dicRowAdd, ncols1, dicColAdd
        #self.tableWidget_2 = QtGui.QTableWidget(self.centralwidget)
        init = QtGui.QTableWidgetSelectionRange(0, 0, nrows1 + len(dicRowAdd) - 1, ncols1 + len(dicColAdd) - 1)
        self.tableWidget_2.setRangeSelected(init, False)

        ranges = QtGui.QTableWidgetSelectionRange(ran[0], ran[1], ran[2], ran[3])
        self.tableWidget_2.setRangeSelected(ranges, True)
        self.tableWidget_2.verticalScrollBar().setSliderPosition(ran[0])
        #self.tableWidget.verticalScrollBar().setSliderPosition(ran[0])

    def select3(self, ran):
        global nrows1, dicRowAdd, ncols1, dicColAdd
        # self.tableWidget_2 = QtGui.QTableWidget(self.centralwidget)
        init = QtGui.QTableWidgetSelectionRange(0, 0, nrows1 + len(dicRowAdd) - 1, ncols1 + len(dicColAdd) - 1)
        ranges = QtGui.QTableWidgetSelectionRange(ran[0], ran[1], ran[2], ran[3])
        self.tableWidget_2.setRangeSelected(init, False)
        self.tableWidget_2.setRangeSelected(ranges, True)
        self.tableWidget_2.verticalScrollBar().setSliderPosition(ran[0])
        self.tableWidget.setRangeSelected(init, False)
        self.tableWidget.setRangeSelected(ranges, True)
        self.tableWidget.verticalScrollBar().setSliderPosition(ran[0])
        # self.tableWidget.verticalScrollBar().setSliderPosition(ran[0])


# 小弹窗
class Ui_aboutWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName(_fromUtf8("MainWindow"))
        MainWindow.resize(368, 181)
        MainWindow.setStyleSheet(_fromUtf8("background-color: rgb(255, 255, 255);"))
        MainWindow.setWindowIcon(QtGui.QIcon(':/ico/Cut.ico'))
        MainWindow.setFixedSize(MainWindow.width(), MainWindow.height())
        self.centralwidget = QtGui.QWidget(MainWindow)
        self.centralwidget.setObjectName(_fromUtf8("centralwidget"))
        self.pushButton = QtGui.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(130, 120, 93, 28))
        self.pushButton.setObjectName(_fromUtf8("pushButton"))
        self.textEdit = QtGui.QTextEdit(self.centralwidget)
        self.textEdit.setGeometry(QtCore.QRect(30, 10, 311, 101))
        self.textEdit.setObjectName(_fromUtf8("textEdit"))
        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtGui.QStatusBar(MainWindow)
        self.statusbar.setObjectName(_fromUtf8("statusbar"))
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        QtCore.QObject.connect(self.pushButton, QtCore.SIGNAL(_fromUtf8("clicked()")),self.Quit)

    def retranslateUi(self, MainWindow):
        MainWindow.setWindowTitle(_translate("MainWindow", "AboutLittleDiff", None))
        self.pushButton.setText(_translate("MainWindow", "OK", None))
        self.textEdit.setHtml(_translate("MainWindow", "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
        "<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
        "p, li { white-space: pre-wrap; }\n"
        "</style></head><body style=\" font-family:\'SimSun\'; font-size:9pt; font-weight:400; font-style:normal;\">\n"
        "<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">Copyright: YuLong Tu </p>\n"
        "<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">Version: 1.0</p>\n"
        "<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">Time: 02/05/2018</p>\n"
        "<p style=\"-qt-paragraph-type:empty; margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><br /></p>\n"
        "<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">Winner, winner, chicken dinner!</p></body></html>", None))

    def Quit(self):
        aboutWindow.close()

# 生成对比表格时，QTableWidget控件无法显示数字，需要把数字转换成str
def FloatAndUnicode(temp):
    if type(temp) == type(1.1) or type(temp) == type(1) or type(temp) == type(1+1j) :
        temp=str(temp)
    else:
        pass
    return temp

def GetSheetRowVal(sheet,num,flag,ncols,nrows):
    #global ncols2
    if flag==0:
        if num>=nrows:
            val1 = []
        else:
            val1=sheet.row_values(num)
        val = []
        for j in val1:
            Data=j
            if type(Data) == type(1L) or type(Data) == type(1):
                val.append(unicode(float(Data)))
            else:
                val.append(unicode(Data))
        return val
    else:
        val=[]
        if num > nrows-1:
            val = []
        else :

            for i in range(1, ncols + 1):
                    Data = sheet.cell(row=num+1, column=i).value
                    if type(Data) == type(1L) or type(Data) == type(1):
                        val.append(unicode(float(Data)))
                    else:
                        val.append(unicode(Data))
        return val
'''
def SheetSpace(sheet,flag):
    rowCol=[-1,-1]
    if flag==0:
        tempr = sheet.nrows
        tempc = sheet.ncols
        for i in range(0, tempr):
            data = sheet.row_values(i)
            for j in range(0,tempc):
                print j
                if (unicode(data[j])==u''or unicode(data[j])==u' ')and j!=0:
                    rowCol[1]=j
                    print 'rowCol[1]',rowCol[1]
                    break
            if rowCol[1]!=-1:
                break
        for i in range(0,tempr):
            data=sheet.row_values(i)
            for j in range(0,tempc):
                if (unicode(data[j])==u''or unicode(data[j])==u' ')and j==0:
                    rowCol[0]=i
                    break
            if rowCol[0]!=-1:
                break

    else:
        tempr = sheet.max_row
        tempc = sheet.max_col
        for i in range(1, tempr + 1):
            for j in range(1, tempc + 1):
                Data = sheet.cell(row=i, column=j).value
                if (unicode(Data)==u'' or unicode(Data)==u' ')and j==0:
                    rowCol[0] = i-1
                    break
        for i in range(1, tempr + 1):
            for j in range(1, tempc + 1):
                Data = sheet.cell(row=i, column=j).value
                if  (unicode(Data)==u'' or unicode(Data)==u' ')and j!=0:
                    rowCol[1] = j-1
                    break
    return rowCol
    '''
if __name__ == "__main__":
    app = QtGui.QApplication(sys.argv)

    # 启动界面
    splash = QtGui.QSplashScreen(QtGui.QPixmap(':/ico/timg.jpg'))
    splash.show()
    app.processEvents()

    # 按钮等控件样式加载
    file = QtCore.QFile('csstest.qss')
    file.open(QtCore.QFile.ReadOnly)
    styleSheet = file.readAll()
    styleSheet = unicode(styleSheet, encoding='utf8')
    QtGui.qApp.setStyleSheet(styleSheet)
    # 登录窗口
    logWindow = QtGui.QMainWindow()
    uiLog = Ui_Log()
    uiLog.setupUi(logWindow)
    # 启动界面
    splash.finish(logWindow)
    logWindow.show()

    #主窗口
    mainWindow = QtGui.QMainWindow()
    uiMain = Ui_MainWindow()
    uiMain.setupUi(mainWindow)
    #mainWindow.show()
    #About窗口
    aboutWindow = QtGui.QMainWindow()
    uiAbout = Ui_aboutWindow()
    uiAbout.setupUi(aboutWindow)
    #aboutWindow.show()

    sys.exit(app.exec_())

